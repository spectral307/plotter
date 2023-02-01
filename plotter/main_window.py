from glob import glob
from os.path import basename, dirname, join
from typing import Union
# Reason: PyQt6 is a third party module.
# pylint: disable-next=no-name-in-module
from PyQt6.QtCore import QSettings
# Reason: PyQt6 is a third party module.
# pylint: disable-next=no-name-in-module
from PyQt6.QtWidgets import QApplication, QFileDialog, QMainWindow
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from .entry_canvas import EntryCanvas
from .gtl_afc_report_entries import GtlAfcReportEntries
from .ui_main_window import Ui_MainWindow


class MainWindow(QMainWindow):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.__ui = Ui_MainWindow()
        self.__ui.setupUi(self)

        self.__ui.open_files.triggered.connect(self.open_files)
        self.__ui.open_folder.triggered.connect(self.open_folder)
        self.__ui.exit_app.triggered.connect(self.exit_app)

        self.__ui.canvas = EntryCanvas()

        self.__ui.splitter.replaceWidget(1, self.__ui.canvas)

    def exit_app(self):
        QApplication.quit()

    # Reason: the method will be changed and decomposed later.
    # pylint: disable-next=too-many-locals
    def open_files(self):
        caption = "Открыть файлы"
        settings = QSettings()
        directory = settings.value("last_dir")
        filefilter = "Все файлы (*.*)"
        files, _ = QFileDialog.getOpenFileNames(
            self, caption, directory, filefilter)
        if not files:
            return

        new_directory = dirname(files[0])
        if new_directory != directory:
            settings.setValue("last_dir", new_directory)

        entries = GtlAfcReportEntries.create(files)
        self.__ui.entries.set_entries(list(entries.keys()))
        self.__ui.canvas.clear()
        for item in list(entries.items()):
            self.__ui.canvas.plot(item[0], item[1]["F"], item[1]["Sabs"])

    # Reason: the method will be changed and decomposed later.
    # pylint: disable-next=too-many-locals
    def open_folder(self):
        caption = "Открыть папку"
        settings = QSettings()
        directory = settings.value("last_dir")
        folder = QFileDialog.getExistingDirectory(
            self, caption, directory)
        if not folder:
            return

        pathname = join(folder, "*.*")
        files = glob(pathname)

        new_directory = dirname(folder)
        if new_directory != directory:
            settings.setValue("last_dir", new_directory)

        entries = GtlAfcReportEntries.create(files)
        self.__ui.entries.set_entries(list(entries.keys()))
        self.__ui.canvas.clear()
        for item in list(entries.items()):
            self.__ui.canvas.plot(item[0], item[1]["F"], item[1]["Sabs"])

    # Reason: the method will be changed and decomposed later.
    # pylint: disable-next=too-many-locals
    def __get_entries(self, files: list[str]):
        entries: dict[str, dict[str, Union[list[Union[int, float]], str]]] = {}

        for file in files:
            entry_name = basename(file)
            datasets: dict[str, Union[list[Union[int, float]], str]] = {}
            entries[entry_name] = datasets

            workbook = load_workbook(file)
            worksheet = workbook["data"]

            x_start_cell = worksheet["E2"]
            sens_abs_start_cell = worksheet["F2"]

            datasets["x"] = self.__read_column(worksheet, x_start_cell)
            datasets["x_label"] = "F, Hz"
            datasets["y"] = self.__read_column(worksheet, sens_abs_start_cell)
            datasets["y_label"] = "Sensitivity abs"

        return entries

    # Reason: the method will be changed and decomposed later.
    # pylint: disable-next=too-many-locals
    def __read_column(self, worksheet: Worksheet, start_cell: Cell):
        i = 0
        start_row = start_cell.row
        col_idx = start_cell.col_idx
        col_data = []
        value = worksheet.cell(start_row, col_idx).value
        while value is not None:
            col_data.append(value)
            i += 1
            value = worksheet.cell(start_row+i, col_idx).value
        return col_data
